import openpyxl


# load the workbook
workbook = openpyxl.load_workbook("RegistraionData.xlsx")

# activate the sheet of the loaded workbook
sheet = workbook.active

# get the cell data
#cell = sheet.cell(row=2,column=2)

# print cell data on console
#print(cell.value)

# write data in the cell
sheet.cell(row=2,column=3).value ="bhavesh1@gmail.com"

# print to check data write in excel cell or not
#print(sheet.cell(row=2,column=3).value)

# find the max row in the sheet
#print(sheet.max_row)

# find the max column in the sheet
#print(sheet.max_column)


# print all data of the sheet

Dict = {}

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == 'Testcase2':
        for j in range(2,sheet.max_column+1):
            #print(sheet.cell(row=i,column=j).value,end=" ")
            # #Dict["firstname"] = "Bhavesh"
            Dict[sheet.cell(row=1,column=j).value] = sheet.cell(row=i, column=j).value
            #print(Dict)

print(Dict)


    #print("")

