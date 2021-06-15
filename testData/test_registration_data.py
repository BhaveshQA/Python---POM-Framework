import openpyxl


class Registation_Data:
    test_registration_data = [{"firstname": "Bhavesh", "email": "bhavesh@mailinator.com"},
                              {"firstname": "Hina", "email": "hina@gamil.com"}]

    @staticmethod
    # static method access using direct class name
    # if method is staticmethod then no need to 'self' variable
    def get_registration_data_excel():#test_case_name
        workbook = openpyxl.load_workbook(r"F:\ProfessionSkillImprovement\PythonSeleniumFramework\testData\RegistraionData.xlsx")
        sheet = workbook.active
        Dict = {}

        for i in range(1, sheet.max_row + 1):
            #if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        # in test we required data
        return [Dict]
